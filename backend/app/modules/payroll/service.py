import csv
import hashlib
import io
import secrets
from datetime import UTC, date, datetime, timedelta
from typing import Any

from fastapi import HTTPException
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.database import parse_uuid
from app.modules.hrd.models import Employee, EmployeeStatus
from app.modules.payroll.models import (
    AttendanceSummary,
    PayrollRun,
    PayrollRunStatus,
    PayrollRunToken,
    PayrollRunType,
    Payslip,
)
from app.modules.payroll.schemas import (
    AttendanceUpsert,
    GenerateSlipsRequest,
    RunCreate,
    TaxPreviewIn,
)
from app.modules.payroll.tax import TaxProfile, compute_pasal17_monthly_average, compute_ter
from app.modules.rates.service import get_effective_bpjs, get_effective_pph21

# Transisi status yang diizinkan per jenis payrol (ADR-0006 / PRD Fase 9).
_ALLOWED_TRANSITIONS: dict[PayrollRunType, dict[PayrollRunStatus, set[PayrollRunStatus]]] = {
    PayrollRunType.internal: {
        PayrollRunStatus.draft: {PayrollRunStatus.finance_processing, PayrollRunStatus.final},
        PayrollRunStatus.finance_processing: {PayrollRunStatus.final},
    },
    PayrollRunType.proyek: {
        PayrollRunStatus.draft: {PayrollRunStatus.submitted_to_client},
        PayrollRunStatus.client_rejected: {PayrollRunStatus.submitted_to_client},
        # Diputuskan klien hanya lewat link ber-token (decide_by_token).
        PayrollRunStatus.submitted_to_client: {
            PayrollRunStatus.client_approved,
            PayrollRunStatus.client_rejected,
        },
        PayrollRunStatus.client_approved: {PayrollRunStatus.finance_processing},
        PayrollRunStatus.finance_processing: {PayrollRunStatus.final},
    },
}

# Status saat slip masih boleh dibuat/diperbarui.
_EDITABLE_STATUSES: dict[PayrollRunType, set[PayrollRunStatus]] = {
    PayrollRunType.internal: {PayrollRunStatus.draft},
    PayrollRunType.proyek: {PayrollRunStatus.draft, PayrollRunStatus.client_rejected},
}


def _transition(run: PayrollRun, target: PayrollRunStatus) -> None:
    allowed = _ALLOWED_TRANSITIONS[run.run_type].get(run.status, set())
    if target not in allowed:
        raise HTTPException(
            status_code=409,
            detail=f"Perubahan status {run.status.value} → {target.value} tidak diizinkan",
        )
    run.status = target


# ---------- Attendance & approval klien ----------


def upsert_attendance(db: Session, payload: AttendanceUpsert) -> AttendanceSummary:
    if db.get(Employee, payload.employee_id) is None:
        raise HTTPException(status_code=404, detail="Karyawan tidak ditemukan")
    summary = db.execute(
        select(AttendanceSummary).where(
            AttendanceSummary.employee_id == payload.employee_id,
            AttendanceSummary.year == payload.year,
            AttendanceSummary.month == payload.month,
        )
    ).scalar_one_or_none()
    if summary is None:
        summary = AttendanceSummary(**payload.model_dump())
        db.add(summary)
    else:
        for field in ("present_days", "overtime_hours", "notes"):
            setattr(summary, field, getattr(payload, field))
        # Perubahan data me-reset approval klien agar diverifikasi ulang.
        summary.client_approved = False
        summary.approved_at = None
    db.commit()
    db.refresh(summary)
    return summary


def set_client_approval(db: Session, attendance_id: str, approved: bool) -> AttendanceSummary:
    summary = db.get(AttendanceSummary, parse_uuid(attendance_id))
    if summary is None:
        raise HTTPException(status_code=404, detail="Rekap absensi tidak ditemukan")
    # Validasi dua jalur (Fase 8): approval klien hanya untuk karyawan eksternal.
    if summary.employee.employment_type.value != "eksternal":
        raise HTTPException(
            status_code=422,
            detail="Karyawan internal divalidasi oleh HR (jalur /attendance/summaries)",
        )
    summary.client_approved = approved
    summary.approved_at = datetime.now(UTC) if approved else None
    db.commit()
    db.refresh(summary)
    return summary


def list_attendance(db: Session, year: int, month: int) -> list[AttendanceSummary]:
    stmt = (
        select(AttendanceSummary)
        .where(AttendanceSummary.year == year, AttendanceSummary.month == month)
        .order_by(AttendanceSummary.created_at)
    )
    return list(db.execute(stmt).scalars())


# ---------- Payroll runs ----------


def _get_run(db: Session, run_id: str) -> PayrollRun:
    run = db.get(PayrollRun, parse_uuid(run_id))
    if run is None:
        raise HTTPException(status_code=404, detail="Payroll run tidak ditemukan")
    return run


def create_run(db: Session, payload: RunCreate) -> PayrollRun:
    if payload.run_type == PayrollRunType.proyek and payload.client_id is not None:
        from app.modules.clients.models import Client

        if db.get(Client, payload.client_id) is None:
            raise HTTPException(status_code=404, detail="Klien tidak ditemukan")
    duplicate = db.execute(
        select(PayrollRun).where(
            PayrollRun.year == payload.year,
            PayrollRun.month == payload.month,
            PayrollRun.run_type == payload.run_type,
            PayrollRun.client_id == payload.client_id,
        )
    ).scalar_one_or_none()
    if duplicate is not None:
        label = f"payrol {payload.run_type.value}"
        if payload.client_id:
            from app.modules.clients.models import Client

            client = db.get(Client, payload.client_id)
            label += f" klien {client.name}" if client else ""
        raise HTTPException(
            status_code=409,
            detail=f"{label.capitalize()} periode {payload.year}-{payload.month} sudah ada",
        )
    run = PayrollRun(
        year=payload.year,
        month=payload.month,
        run_type=payload.run_type,
        client_id=payload.client_id,
    )
    db.add(run)
    db.commit()
    db.refresh(run)
    return run


def list_runs(db: Session) -> list[PayrollRun]:
    stmt = select(PayrollRun).order_by(PayrollRun.year.desc(), PayrollRun.month.desc())
    return list(db.execute(stmt).scalars())


def get_run(db: Session, run_id: str) -> PayrollRun:
    return _get_run(db, run_id)


def list_slips(db: Session, run_id: str) -> list[Payslip]:
    run = _get_run(db, run_id)
    return list(run.slips)


def generate_slips(db: Session, run_id: str, payload: GenerateSlipsRequest) -> list[Payslip]:
    """Buat slip gaji untuk karyawan aktif; hanya lembur yang disetujui klien."""
    run = _get_run(db, run_id)
    if run.status not in _EDITABLE_STATUSES[run.run_type]:
        raise HTTPException(
            status_code=409,
            detail=(
                "Slip hanya bisa dibuat saat status draft (atau ditolak klien untuk payrol proyek)"
            ),
        )

    # Resolve rate ber-versi untuk periode ini; snapshot disimpan di run untuk historis.
    period_date = date(run.year, run.month, 1)
    pph21_cfg = get_effective_pph21(db, period_date)
    bpjs_cfg = get_effective_bpjs(db, period_date)
    # Simpan snapshot JSON agar laporan historis tetap konsisten walau rate diperbarui
    if pph21_cfg and run.pph21_snapshot is None:
        run.pph21_snapshot = {
            "effective_from": pph21_cfg.effective_from.isoformat(),
            "ptkp_diri": float(pph21_cfg.ptkp_diri),
            "config_id": pph21_cfg.id,
        }
    if bpjs_cfg and run.bpjs_snapshot is None:
        run.bpjs_snapshot = {
            "effective_from": bpjs_cfg.effective_from.isoformat(),
            "config_id": bpjs_cfg.id,
        }

    # Fase 26 butir 4: employee.payroll_locked -- karyawan terkunci dilewati
    # saat generate slip baru (bukan cuma cegah edit slip yang sudah ada).
    employee_stmt = select(Employee).where(
        Employee.status == EmployeeStatus.active, Employee.payroll_locked.is_(False)
    )
    if payload.employee_ids:
        ids = [parse_uuid(str(e)) for e in payload.employee_ids]
        employee_stmt = employee_stmt.where(Employee.id.in_(ids))
    # Payrol proyek: hanya karyawan yang ditempatkan di klien run ini.
    if run.run_type == PayrollRunType.proyek and run.client_id is not None:
        from app.modules.clients.models import Client  # noqa: F401
        from app.modules.recruitment.models import JobOrder, Placement

        employee_stmt = (
            employee_stmt.join(Placement, Employee.placement_id == Placement.id)
            .join(JobOrder, Placement.job_order_id == JobOrder.id)
            .where(JobOrder.client_id == run.client_id)
        )
    employees = list(db.execute(employee_stmt).scalars())
    if not employees:
        detail = "Tidak ada karyawan aktif untuk diproses"
        if run.run_type == PayrollRunType.proyek:
            detail = "Tidak ada karyawan aktif yang ditempatkan di klien ini"
        raise HTTPException(status_code=422, detail=detail)

    summaries = {
        s.employee_id: s
        for s in db.execute(
            select(AttendanceSummary).where(
                AttendanceSummary.year == run.year, AttendanceSummary.month == run.month
            )
        ).scalars()
    }

    # Hari kerja bulan tsb (Sen–Jum) untuk prorata absensi.
    workdays = _weekday_count(run.year, run.month)

    slips: list[Payslip] = []
    skipped_unapproved = 0
    for employee in employees:
        overtime_hours = 0
        summary = summaries.get(employee.id)
        if summary is not None and summary.overtime_hours:
            if not summary.client_approved:
                skipped_unapproved += 1
            else:
                overtime_hours = summary.overtime_hours

        base_full = float(employee.base_salary or 0)
        allowance_full = float(payload.allowance or 0)
        overtime_amount = overtime_hours * float(payload.overtime_rate or 0)

        # Prorata opt-in (PRD §6.1): dari rekap absensi TERVALIDASI.
        eff_ratio = 1.0
        effective_days: int | None = None
        if payload.prorata_absensi and summary is not None:
            effective_days = max(summary.present_days, 0)
            if workdays > 0 and effective_days < workdays:
                eff_ratio = effective_days / workdays
        base = round(base_full * eff_ratio)
        allowance_amt = round(allowance_full * eff_ratio)
        gross = base + allowance_amt + overtime_amount

        # Gunakan config ber-versi jika ada, fallback ke konstanta kode
        profile = TaxProfile.from_db(
            db,
            period_date,
            marital_status=(employee.marital_status.value if employee.marital_status else "tk"),
            dependents=employee.dependents or 0,
        )
        tax = compute_ter(gross, profile)
        # Potongan admin bank otomatis (non-Mandiri) dari config
        try:
            from app.modules.rates.service import get_bank_fee

            bank_fee = get_bank_fee(db, employee.bank_name or "")
        except Exception:
            bank_fee = 0

        # BPJS dua sisi (opt-in): potongan karyawan + passthrough perusahaan
        bpjs_emp_total = 0
        breakdown = None
        if payload.bpjs_enabled:
            from app.modules.bpjs.engine import compute_contribution

            breakdown = compute_contribution(
                base,
                jkk_risk_category=employee.jkk_risk_category,
                db=db,
                effective_date=period_date,
            )
            bpjs_emp_total = breakdown.kes_employee + breakdown.jht_employee + breakdown.jp_employee

        total_deductions = float(payload.deductions or 0) + bank_fee + bpjs_emp_total
        slip = Payslip(
            run_id=run.id,
            employee_id=employee.id,
            base_salary=base,
            allowance=allowance_amt,
            overtime_hours=overtime_hours,
            overtime_rate=float(payload.overtime_rate or 0),
            overtime_amount=overtime_amount,
            deductions=total_deductions,
            gross=gross,
            pph21_method="ter",
            tax_pph21=tax,
            net_pay=gross - tax - total_deductions,
        )

        # Line-item Saltab (PRD §6) — dari angka yang sama agar komponen ↔
        # agregat slip selalu "nol selisih".
        prorata_note = (
            f"Prorata {effective_days}/{workdays} hari kerja"
            if effective_days is not None and eff_ratio < 1.0
            else None
        )
        comps_spec: list[tuple[str, str, str, float]] = [
            ("earnings", "gaji_pokok", "Gaji pokok", base),
            ("earnings", "tunjangan", "Tunjangan", allowance_amt),
        ]
        if overtime_amount:
            comps_spec.append(("earnings", "lembur", "Lembur", overtime_amount))
        comps_spec.append(("deduction", "pph21", "PPh 21", tax))
        if payload.deductions:
            comps_spec.append(
                ("deduction", "potongan_lain", "Potongan lain", float(payload.deductions))
            )
        if bank_fee:
            comps_spec.append(("deduction", "admin_bank", "Admin bank", bank_fee))
        if breakdown is not None:
            comps_spec.extend(
                [
                    (
                        "deduction",
                        "bpjs_kesehatan_py",
                        "BPJS Kesehatan (karyawan)",
                        breakdown.kes_employee,
                    ),
                    ("deduction", "jht_py", "JHT (karyawan)", breakdown.jht_employee),
                    ("deduction", "jp_py", "JP (karyawan)", breakdown.jp_employee),
                    (
                        "passthrough",
                        "bpjs_employer",
                        "BPJS tanggungan perusahaan",
                        breakdown.employer_total,
                    ),
                ]
            )
        slip._saltab_comps = [  # type: ignore[attr-defined]  # noqa: SLF001
            {
                "ctype": c,
                "code": k,
                "name": n,
                "amount": a,
                "notes": prorata_note if k == "gaji_pokok" else None,
            }
            for c, k, n, a in comps_spec
            if a > 0
        ]  # type: ignore[attr-defined]
        slips.append(slip)

    existing = {
        (s.run_id, s.employee_id)
        for s in db.execute(select(Payslip).where(Payslip.run_id == run.id)).scalars()
    }
    added = [s for s in slips if (s.run_id, s.employee_id) not in existing]
    if not added:
        message = "Semua slip sudah ada"
        if skipped_unapproved:
            message += f"; {skipped_unapproved} lembur belum disetujui klien"
        raise HTTPException(status_code=409, detail=message)
    db.add_all(added)
    db.flush()
    # Simpan line-item Saltab untuk slip baru.
    from app.modules.payroll.models import PayslipComponent

    for slip in added:
        spec = getattr(slip, "_saltab_comps", None) or []
        db.add_all(
            [
                PayslipComponent(
                    payslip_id=slip.id,
                    ctype=item["ctype"],
                    code=item["code"],
                    name=item["name"],
                    amount=item["amount"],
                    source="auto",
                    notes=item.get("notes"),
                )
                for item in spec
            ]
        )
    db.commit()
    for slip in added:
        db.refresh(slip)
    return added


def _weekday_count(year: int, month: int) -> int:
    """Hari kerja Sen–Jum dalam satu bulan (basis prorata)."""
    from calendar import monthrange

    last = monthrange(year, month)[1]
    start = date(year, month, 1)
    end = date(year, month, last)
    days = 0
    cur = start
    while cur <= end:
        if cur.weekday() < 5:
            days += 1
        cur = date.fromordinal(cur.toordinal() + 1)
    return days


# ---------- Saltab grid (PRD §6.3) ----------


def saltab_view(db: Session, run_id: str, tenant_id=None) -> list[dict]:
    run = _get_run(db, run_id)
    rows: list[dict] = []
    for slip in sorted(run.slips, key=lambda s: s.employee.full_name):
        comps: list[dict[str, Any]] = [
            {
                "id": str(c.id),
                "ctype": c.ctype.value,
                "code": c.code,
                "name": c.name,
                "amount": float(c.amount),
                "source": c.source,
                "notes": c.notes,
            }
            for c in slip.components
        ]
        total_earnings = sum(c["amount"] for c in comps if c["ctype"] == "earnings")
        total_deductions = sum(c["amount"] for c in comps if c["ctype"] == "deduction")
        total_passthrough = sum(c["amount"] for c in comps if c["ctype"] == "passthrough")
        rows.append(
            {
                "payslip_id": str(slip.id),
                "employee_id": str(slip.employee_id),
                "employee_name": slip.employee.full_name,
                "status_run": run.status.value,
                "components": comps,
                "total_earnings": round(total_earnings),
                "total_deductions": round(total_deductions),
                "total_passthrough": round(total_passthrough),
            }
        )
    return rows


def update_saltab_component(db: Session, user, component_id: str, amount: float):
    """Override manual komponen (grid Saltab); agregat slip dihitung ulang."""
    from app.modules.payroll.models import PayslipComponent, PayslipComponentType

    comp = db.get(PayslipComponent, parse_uuid(component_id))
    if comp is None:
        raise HTTPException(status_code=404, detail="Komponen tidak ditemukan")
    slip = db.get(Payslip, comp.payslip_id)
    if slip is None:
        raise HTTPException(status_code=404, detail="Slip tidak ditemukan")
    run = _get_run(db, str(slip.run_id))
    if run.status not in _EDITABLE_STATUSES[run.run_type]:
        raise HTTPException(
            status_code=409, detail="Grid hanya bisa diedit saat draft/ditolak klien"
        )
    employee = db.get(Employee, slip.employee_id)
    if employee is not None and employee.payroll_locked:
        raise HTTPException(status_code=409, detail="Payroll karyawan ini terkunci")

    old = float(comp.amount)
    comp.amount = amount
    comp.source = "manual"
    comp.notes = f"Override manual oleh {user.email}"

    earnings_total = 0.0
    deductions_excl_tax = 0.0
    tax_amount = float(slip.tax_pph21)
    for c in slip.components:
        amt = float(c.amount)
        if c.ctype == PayslipComponentType.earnings:
            earnings_total += amt
        elif c.code == "pph21":
            tax_amount = amt
        elif c.ctype == PayslipComponentType.deduction:
            deductions_excl_tax += amt

    slip.gross = round(earnings_total)
    slip.deductions = round(deductions_excl_tax)
    slip.net_pay = round(earnings_total) - round(tax_amount) - round(deductions_excl_tax)

    db.commit()
    db.refresh(slip)
    from app.modules import audit

    audit.log_event(
        db,
        action="saltab.component_override",
        entity_type="payslip_component",
        entity_id=comp.id,
        detail={
            "old": old,
            "new": amount,
            "employee_id": str(slip.employee_id),
            "by": user.email,
        },
    )
    return comp, slip


def saltab_export_csv(db: Session, run_id: str) -> tuple[str, str]:
    """Ekspor CSV grid Saltab (pengganti file Excel manual)."""
    run = _get_run(db, run_id)
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer, delimiter=";")
    writer.writerow(["Karyawan", "Komponen", "Jenis", "Nominal", "Sumber"])
    for row in saltab_view(db, run_id):
        for comp in row["components"]:
            writer.writerow(
                [
                    row["employee_name"],
                    comp["name"],
                    comp["ctype"],
                    comp["amount"],
                    comp["source"],
                ]
            )
        thp = row["total_earnings"] - row["total_deductions"]
        writer.writerow([row["employee_name"], "TOTAL THP", "", thp, ""])
        writer.writerow([])
    filename = f"saltab-{run.year}{run.month:02d}-{str(run.id)[:8]}.csv"
    return buffer.getvalue(), filename


def saltab_export_excel(db: Session, run_id: str) -> tuple[bytes, str]:
    """Ekspor Excel Saltab dengan styling header."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    run = _get_run(db, run_id)
    wb = Workbook()
    ws = wb.active
    ws.title = f"Saltab {run.year}-{run.month:02d}"

    header = ["Karyawan", "Komponen", "Jenis", "Nominal", "Sumber"]
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    ws.append(header)
    for col in range(1, 6):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = align_center
        c.border = border
    ws.row_dimensions[1].height = 18

    thp_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    thp_font = Font(bold=True)

    row_idx = 2
    for row in saltab_view(db, run_id):
        for comp in row["components"]:
            ws.append(
                [row["employee_name"], comp["name"], comp["ctype"], comp["amount"], comp["source"]]
            )  # noqa: E501
            for col, align in [
                (1, align_left),
                (2, align_left),
                (3, align_center),
                (4, align_right),
                (5, align_center),
            ]:  # noqa: E501
                c = ws.cell(row=row_idx, column=col)
                c.border = border
                c.alignment = align
                if col == 4:
                    c.number_format = "#,##0"
            row_idx += 1
        thp = row["total_earnings"] - row["total_deductions"]
        ws.append([row["employee_name"], "TOTAL THP", "", thp, ""])
        for col in range(1, 6):
            c = ws.cell(row=row_idx, column=col)
            c.fill = thp_fill
            c.font = thp_font
            c.border = border
            c.alignment = align_right if col == 4 else align_left
            if col == 4:
                c.number_format = "#,##0"
        row_idx += 1
        ws.append([])
        row_idx += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.print_title_rows = "1:1"

    buf = io.BytesIO()
    wb.save(buf)
    filename = f"saltab-{run.year}{run.month:02d}-{str(run.id)[:8]}.xlsx"
    return buf.getvalue(), filename


def saltab_export_pdf(db: Session, run_id: str) -> tuple[bytes, str]:
    """Ekspor PDF Saltab via reportlab."""
    from io import BytesIO

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    run = _get_run(db, run_id)
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=f"Saltab {run.year}-{run.month:02d}",
    )
    styles = getSampleStyleSheet()
    title_style = styles["Heading2"]
    title_style.textColor = colors.HexColor("#1F3864")
    normal = styles["Normal"]
    normal.fontSize = 8

    story = []
    story.append(
        Paragraph(
            f"Saltab &mdash; {run.year}/{run.month:02d} &mdash; Run {str(run.id)[:8]}", title_style
        )
    )  # noqa: E501
    story.append(Spacer(1, 6))

    for row in saltab_view(db, run_id):
        data = [["Karyawan", "Komponen", "Jenis", "Nominal", "Sumber"]]
        for comp in row["components"]:
            data.append(
                [
                    row["employee_name"],
                    comp["name"],
                    comp["ctype"],
                    f"{comp['amount']:,.0f}",
                    comp["source"],
                ]
            )  # noqa: E501
        thp = row["total_earnings"] - row["total_deductions"]
        data.append([row["employee_name"], "TOTAL THP", "", f"{thp:,.0f}", ""])
        tbl = Table(data, colWidths=[45 * mm, 55 * mm, 28 * mm, 30 * mm, 28 * mm], repeatRows=1)
        tbl.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F3864")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7.5),
                    ("ALIGN", (3, 0), (3, -1), "RIGHT"),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D9D9D9")),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor("#F2F2F2")],
                    ),  # noqa: E501
                    ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#FFF2CC")),
                    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        # wrap text in cells
        for r_idx in range(len(data)):
            for c_idx in range(len(data[0])):
                if isinstance(data[r_idx][c_idx], str) and len(data[r_idx][c_idx]) > 28:
                    data[r_idx][c_idx] = Paragraph(data[r_idx][c_idx], normal)
        story.append(Paragraph(f"<b>{row['employee_name']}</b>", normal))
        story.append(Spacer(1, 2))
        story.append(tbl)
        story.append(Spacer(1, 10))

    doc.build(story)
    filename = f"saltab-{run.year}{run.month:02d}-{str(run.id)[:8]}.pdf"
    return buf.getvalue(), filename


def send_saltab_to_client(db: Session, run_id: str, recipient_email: str) -> None:
    """Kirim manual Saltab (PDF) ke email klien -- Fase 23 butir 4.

    Tombol ini dipencet Ops, BUKAN dikirim otomatis begitu link dibuat, dan
    email penerima diisi manual tiap kirim (bukan diambil dari data klien)
    karena PIC bisa beda-beda per pengiriman -- keputusan eksplisit PRD."""
    from app.modules import audit

    run = _get_run(db, run_id)
    content, filename = saltab_export_pdf(db, run_id)
    from app.modules.notifications.service import send_raw_email_with_attachment

    send_raw_email_with_attachment(
        recipient_email,
        f"Saltab {run.year}-{run.month:02d}",
        "Terlampir rekap Saltab (slip gaji tabel) untuk periode ini.",
        attachment_bytes=content,
        attachment_filename=filename,
        attachment_maintype="application",
        attachment_subtype="pdf",
    )
    audit.log_event(
        db,
        action="payroll.saltab_sent_to_client",
        entity_type="payroll_run",
        entity_id=run.id,
        detail={"recipient_email": recipient_email},
    )


def bukti_potong_pdf(db: Session, run_id: str, employee_id: str) -> tuple[bytes, str]:
    """PRD v3.0 §6 — Bukti Pemotongan PPh 21 per karyawan per run payrol.

    Nomor bukti `1.1-YYYYMM-SEQ`: SEQ berurutan per run diambil dari urutan
    nama karyawan (deterministik, tidak tergantung urutan insert slip).
    """
    from io import BytesIO

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    from app.modules.payroll.tax import TaxProfile, ter_category
    from app.modules.platform.models import Tenant

    run = _get_run(db, run_id)
    emp_uuid = parse_uuid(employee_id)
    ordered_slips = sorted(
        run.slips, key=lambda s: (s.employee.full_name or "", str(s.employee_id))
    )
    seq = None
    slip = None
    for idx, s in enumerate(ordered_slips, start=1):
        if s.employee_id == emp_uuid:
            slip = s
            seq = idx
            break
    if slip is None:
        raise HTTPException(status_code=404, detail="Slip karyawan tidak ditemukan pada run ini")

    employee = slip.employee
    tenant = db.get(Tenant, run.tenant_id)
    no_bukti = f"1.1-{run.year}{run.month:02d}-{seq:04d}"

    period_date = date(run.year, run.month, 1)
    profile = TaxProfile.from_db(
        db,
        period_date,
        marital_status=(employee.marital_status.value if employee.marital_status else "tk"),
        dependents=employee.dependents or 0,
    )
    ptkp_label = f"{profile.marital_status.upper()}/{min(employee.dependents or 0, 3)}"

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=20 * mm,
        rightMargin=20 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        title=f"Bukti Potong PPh21 {no_bukti}",
    )
    styles = getSampleStyleSheet()
    title_style = styles["Heading2"]
    title_style.alignment = 1
    normal = styles["Normal"]
    bold = styles["Normal"].clone("bold")
    bold.fontName = "Helvetica-Bold"

    story = [
        Paragraph("BUKTI PEMOTONGAN PAJAK PENGHASILAN PASAL 21", title_style),
        Paragraph("(Pegawai Tetap &mdash; Masa Bulanan)", styles["Italic"]),
        Spacer(1, 4),
        Paragraph(f"No. Bukti Potong: <b>{no_bukti}</b>", normal),
        Paragraph(f"Masa Pajak: {run.month:02d}/{run.year}", normal),
        Spacer(1, 10),
        Paragraph("A. Identitas Pemotong Pajak", bold),
        Table(
            [
                ["Nama", ": " + (tenant.name if tenant else "-")],
                ["NPWP", ": -"],
            ],
            colWidths=[35 * mm, 120 * mm],
        ),
        Spacer(1, 8),
        Paragraph("B. Identitas Penerima Penghasilan", bold),
        Table(
            [
                ["Nama", ": " + employee.full_name],
                ["No. Induk Karyawan", ": " + employee.employee_no],
                ["NPWP", ": " + (employee.npwp_no or "-")],
                ["Status PTKP", ": " + ptkp_label],
            ],
            colWidths=[35 * mm, 120 * mm],
        ),
        Spacer(1, 8),
        Paragraph("C. Rincian Penghasilan &amp; Pemotongan Pajak", bold),
    ]

    rincian = [
        ["Uraian", "Jumlah (Rp)"],
        ["Penghasilan Bruto Sebulan", f"{float(slip.gross):,.0f}"],
        ["PTKP Setahun", f"{profile.ptkp_annual:,.0f}"],
        ["Metode Perhitungan", f"TER Kategori {ter_category(profile)}"],
        ["PPh 21 Dipotong Bulan Ini", f"{float(slip.tax_pph21):,.0f}"],
    ]
    tbl = Table(rincian, colWidths=[100 * mm, 55 * mm])
    tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F3864")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D9D9D9")),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#FFF2CC")),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(tbl)
    story.append(Spacer(1, 24))
    story.append(
        Paragraph(
            f"Dicetak sistem pada {datetime.now(UTC).strftime('%d %B %Y')}. "
            "Dokumen ini sah tanpa tanda tangan basah sesuai kebijakan internal pemotong pajak.",
            styles["Italic"],
        )
    )

    doc.build(story)
    filename = f"bukti-potong-{no_bukti}-{employee.employee_no}.pdf"
    return buf.getvalue(), filename


def _get_employee_slip(db: Session, run_id: str, employee_id: str) -> Payslip:
    run = _get_run(db, run_id)
    slip = db.execute(
        select(Payslip).where(
            Payslip.run_id == run.id, Payslip.employee_id == parse_uuid(employee_id)
        )
    ).scalar_one_or_none()
    if slip is None:
        raise HTTPException(status_code=404, detail="Slip karyawan tidak ditemukan pada run ini")
    return slip


def employee_payslip_pdf(db: Session, run_id: str, employee_id: str) -> tuple[bytes, str]:
    """Payslip lengkap per karyawan -- Fase 26 butir 5, BEDA dari
    `bukti_potong_pdf` (itu sertifikat pajak PPh21, bukan slip gaji utuh)."""
    from io import BytesIO

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    slip = _get_employee_slip(db, run_id, employee_id)
    run = _get_run(db, run_id)
    employee = slip.employee

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=20 * mm,
        rightMargin=20 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        title=f"Payslip {run.year}-{run.month:02d}",
    )
    styles = getSampleStyleSheet()
    title_style = styles["Heading2"]
    normal = styles["Normal"]

    rows = [["Komponen", "Jenis", "Nominal (Rp)"]]
    for comp in slip.components:
        rows.append([comp.name, comp.ctype.value, f"{float(comp.amount):,.0f}"])
    rows.append(["THP (Net Pay)", "", f"{float(slip.net_pay):,.0f}"])

    tbl = Table(rows, colWidths=[80 * mm, 40 * mm, 45 * mm])
    tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F3864")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (2, 0), (2, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D9D9D9")),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#FFF2CC")),
            ]
        )
    )

    story = [
        Paragraph(f"SLIP GAJI &mdash; {run.year}/{run.month:02d}", title_style),
        Spacer(1, 4),
        Paragraph(f"Nama: {employee.full_name}", normal),
        Paragraph(f"No. Induk Karyawan: {employee.employee_no}", normal),
        Spacer(1, 10),
        tbl,
    ]
    doc.build(story)
    filename = f"payslip-{run.year}{run.month:02d}-{employee.employee_no}.pdf"
    return buf.getvalue(), filename


def send_payslip_email(db: Session, run_id: str, employee_id: str) -> None:
    """Kirim payslip langsung ke email karyawan -- Fase 26 butir 5, TERPISAH
    dari alur Ops->klien Fase 23 butir 4 (`send_saltab_to_client`, penerima
    beda: karyawan sendiri, bukan PIC klien)."""
    from app.modules.auth.models import User
    from app.modules.notifications.service import send_raw_email_with_attachment

    slip = _get_employee_slip(db, run_id, employee_id)
    employee = slip.employee
    if employee.user_id is None:
        raise HTTPException(
            status_code=400, detail="Karyawan ini belum tertaut akun (tidak ada email tujuan)"
        )
    user = db.get(User, employee.user_id)
    if user is None or not user.email:
        raise HTTPException(status_code=400, detail="Akun karyawan tidak punya email")

    run = _get_run(db, run_id)
    content, filename = employee_payslip_pdf(db, run_id, employee_id)
    send_raw_email_with_attachment(
        user.email,
        f"Slip Gaji {run.year}-{run.month:02d}",
        "Terlampir slip gaji Anda untuk periode ini.",
        attachment_bytes=content,
        attachment_filename=filename,
        attachment_maintype="application",
        attachment_subtype="pdf",
    )
    from app.modules import audit

    audit.log_event(
        db,
        action="payroll.payslip_emailed",
        entity_type="payslip",
        entity_id=slip.id,
        detail={"employee_id": str(employee.id)},
    )


def finalize_run(db: Session, run_id: str) -> PayrollRun:
    run = _get_run(db, run_id)
    if not run.slips:
        raise HTTPException(status_code=422, detail="Belum ada slip gaji untuk difinalisasi")
    _transition(run, PayrollRunStatus.final)
    run.finalized_at = datetime.now(UTC)
    db.commit()
    db.refresh(run)
    from app.modules import audit

    audit.log_event(
        db,
        action="payroll.finalized",
        entity_type="payroll_run",
        entity_id=run.id,
        detail={"run_type": run.run_type.value, "period": f"{run.year}-{run.month:02d}"},
    )
    # Fase 10: jurnal otomatis payroll_finalized_{internal|proyek} (best-effort).
    try:
        from app.modules.accounting.service import post_auto_event

        gross_total = sum(float(s.gross) for s in run.slips)
        net_total = sum(float(s.net_pay) for s in run.slips)
        tax_total = sum(float(s.tax_pph21) for s in run.slips)
        bpjs_er = sum(
            float(c.amount) for s in run.slips for c in s.components if c.code == "bpjs_employer"
        )
        bpjs_emp = sum(
            float(c.amount)
            for s in run.slips
            for c in s.components
            if c.code in ("bpjs_kesehatan_py", "jht_py", "jp_py")
        )
        event = (
            "payroll_finalized_internal"
            if run.run_type == PayrollRunType.internal
            else "payroll_finalized_proyek"
        )
        lines: list[tuple[str, float, float]] = [("5-1000", round(gross_total), 0.0)]
        if bpjs_er:
            lines.append(("5-3000", round(bpjs_er), 0.0))
        lines.append(("2-1000", 0.0, round(net_total)))
        if tax_total:
            lines.append(("2-1100", 0.0, round(tax_total)))
        if bpjs_emp + bpjs_er:
            lines.append(("2-1200", 0.0, round(bpjs_emp + bpjs_er)))
        post_auto_event(
            db,
            tenant_id=run.tenant_id,
            event_code=event,
            source_ref_type="payroll_run",
            source_ref_id=run.id,
            entry_date=date.today(),
            description=f"Payrol {run.run_type.value} {run.month}/{run.year}",
            lines=lines,
            client_dim_id=run.client_id,
        )
    except Exception:  # noqa: BLE001 - jurnal tidak boleh memblokir bisnis
        import logging

        logging.getLogger(__name__).exception("Auto-journal payrol gagal")
    return run


# ---------- Alur approval klien (link ber-token) ----------


def _hash_token(raw: str) -> str:
    return hashlib.sha256(raw.encode()).hexdigest()


def submit_to_client(db: Session, run_id: str, days: int = 14) -> tuple[PayrollRun, str, datetime]:
    """Kirim payrol proyek ke klien: status berubah + buat link ber-token."""
    run = _get_run(db, run_id)
    if run.run_type != PayrollRunType.proyek:
        raise HTTPException(
            status_code=422,
            detail=(
                "Hanya payrol proyek yang dikirim ke klien; "
                "payrol internal langsung diproses Finance"
            ),
        )
    if not 1 <= days <= 90:
        raise HTTPException(status_code=422, detail="Masa berlaku token 1-90 hari")
    if not run.slips:
        raise HTTPException(status_code=422, detail="Belum ada slip gaji untuk dikirim ke klien")

    # Cabut token lama yang belum terpakai.
    for stale in db.execute(
        select(PayrollRunToken).where(
            PayrollRunToken.run_id == run.id, PayrollRunToken.decided_at.is_(None)
        )
    ).scalars():
        db.delete(stale)

    raw = secrets.token_urlsafe(24)
    expires_at = datetime.now(UTC) + timedelta(days=days)
    db.add(
        PayrollRunToken(
            run_id=run.id,
            token_hash=_hash_token(raw),
            expires_at=expires_at,
        )
    )
    _transition(run, PayrollRunStatus.submitted_to_client)
    db.commit()
    db.refresh(run)
    from app.modules import audit

    audit.log_event(
        db,
        action="payroll.submitted_to_client",
        entity_type="payroll_run",
        entity_id=run.id,
        detail={"expires_days": days, "slips": len(run.slips)},
    )
    try:
        from app.modules.chat.service import post_payroll_status_message

        post_payroll_status_message(
            db, run, f"Payrol {run.month}/{run.year} dikirim menunggu persetujuan klien"
        )
    except Exception:
        pass
    return run, raw, expires_at


def _find_token(db: Session, raw_token: str) -> PayrollRunToken:
    token = db.execute(
        select(PayrollRunToken).where(PayrollRunToken.token_hash == _hash_token(raw_token))
    ).scalar_one_or_none()
    if token is None:
        raise HTTPException(status_code=404, detail="Link approval tidak valid")
    if token.decided_at is not None:
        raise HTTPException(status_code=409, detail="Keputusan untuk link ini sudah direkam")
    expires = token.expires_at
    now = datetime.now(UTC)
    if expires.tzinfo is None:
        now = now.replace(tzinfo=None)
    if expires < now:
        raise HTTPException(status_code=410, detail="Link approval sudah kedaluwarsa")
    return token


def client_view(db: Session, raw_token: str) -> dict:
    """Ringkasan Saltab read-only untuk link approval klien (tanpa akun)."""
    token = _find_token(db, raw_token)
    run = db.get(PayrollRun, token.run_id)
    if run is None:
        raise HTTPException(status_code=404, detail="Payrol tidak ditemukan")
    lines = [
        {
            "employee_name": s.employee.full_name,
            "base_salary": float(s.base_salary),
            "allowance": float(s.allowance),
            "overtime_amount": float(s.overtime_amount),
            "deductions": float(s.deductions),
            "gross": float(s.gross),
            "tax_pph21": float(s.tax_pph21),
            "net_pay": float(s.net_pay),
        }
        for s in run.slips
    ]
    return {
        "client": run.client.name if run.client else None,
        "year": run.year,
        "month": run.month,
        "status": run.status.value,
        "expires_at": token.expires_at,
        "decided": token.decided_at is not None,
        "decided_by_name": token.decided_by_name,
        "decision_note": token.decision_note,
        "lines": lines,
        "total_net_pay": sum(line["net_pay"] for line in lines),
        "total_gross": sum(line["gross"] for line in lines),
    }


def decide_by_token(
    db: Session, raw_token: str, approved: bool, name: str, note: str | None
) -> dict:
    """Rekam keputusan klien dari link publik; transisi status divalidasi.

    Endpoint ini TANPA konteks tenant (publik), sehingga seluruh penulisan
    dibungkus set_tenant(run.tenant_id) agar baris TenantMixin valid.
    """
    from app.core.tenancy import get_tenant, set_tenant

    token = _find_token(db, raw_token)
    run = db.get(PayrollRun, token.run_id)
    if run is None:
        raise HTTPException(status_code=404, detail="Payrol tidak ditemukan")

    prev_tenant = get_tenant()
    set_tenant(run.tenant_id)
    try:
        target = PayrollRunStatus.client_approved if approved else PayrollRunStatus.client_rejected
        _transition(run, target)
        token.decided_at = datetime.now(UTC)
        token.decided_by_name = name[:255]
        token.decision_note = (note or "").strip()[:500] or None
        db.commit()
        db.refresh(run)

        from app.modules import audit
        from app.modules.notifications.service import notify_hr_users

        audit.log_event(
            db,
            action="payroll.client_decision",
            entity_type="payroll_run",
            entity_id=run.id,
            detail={"approved": approved, "by": name},
        )
        keputusan = "disetujui" if approved else "ditolak"
        notify_hr_users(
            db,
            title=f"Payrol proyek {run.month}/{run.year}: {keputusan} klien",
            body=(f"Keputusan oleh {name}" + (f" — {note}" if note else "")),
            entity_id=run.id,
        )
        try:
            from app.modules.chat.service import post_payroll_status_message

            post_payroll_status_message(db, run, f"Payrol {keputusan} klien: {name}")
        except Exception:
            pass

        # Fase 9b: draft invoice otomatis saat klien menyetujui (best-effort).
        if approved and run.client_id is not None:
            try:
                from app.modules.finance import service as finance_service
                from app.modules.finance.schemas import InvoiceGenerateRequest

                finance_service.generate_invoice(
                    db,
                    InvoiceGenerateRequest(
                        client_id=run.client_id,
                        year=run.year,
                        month=run.month,
                        fee_amount=0,
                        notes=f"Otomatis dari payrol proyek ({name})",
                        run_id=run.id,
                    ),
                    run_id=run.id,
                )
            except HTTPException as exc:
                if exc.status_code != 409:  # 409 = invoice periode sudah ada
                    import logging

                    logging.getLogger(__name__).warning(
                        "Invoice otomatis gagal untuk run %s: %s", run.id, exc.detail
                    )
    finally:
        set_tenant(prev_tenant)

    return {
        "status": run.status.value,
        "decided_by_name": token.decided_by_name,
        "decision_note": token.decision_note,
    }


def start_finance_processing(db: Session, run_id: str) -> PayrollRun:
    run = _get_run(db, run_id)
    _transition(run, PayrollRunStatus.finance_processing)
    db.commit()
    db.refresh(run)
    from app.modules import audit

    audit.log_event(
        db,
        action="payroll.finance_processing",
        entity_type="payroll_run",
        entity_id=run.id,
        detail={"run_type": run.run_type.value},
    )
    return run


def resubmit_allowed(run: PayrollRun) -> bool:
    """Setelah ditolak klien, angka boleh diperbaiki lalu dikirim ulang."""
    return run.status == PayrollRunStatus.client_rejected


# ---------- Preview pajak ----------


def preview_tax(payload: TaxPreviewIn, db: Session | None = None) -> dict:
    # Gunakan config ber-versi jika db tersedia
    effective = date.today()
    if db is not None:
        profile = TaxProfile.from_db(db, effective, payload.marital_status, payload.dependents)
    else:
        profile = TaxProfile(marital_status=payload.marital_status, dependents=payload.dependents)
    if payload.method == "pasal17":
        tax = compute_pasal17_monthly_average(payload.gross_monthly, payload.months, profile)
    else:
        tax = compute_ter(payload.gross_monthly, profile)
    return {"tax_pph21": tax, "method": payload.method}
